import os
import sqlite3
import io
from datetime import datetime
from flask import Flask, render_template, request, redirect, url_for, flash, abort, g, jsonify, Response
from dotenv import load_dotenv

load_dotenv()

ADMIN_KEY = os.getenv("ADMIN_KEY", "cambiame-por-una-clave-secreta")
DB_PATH = os.getenv("DB_PATH", "/data/rsvps.db")
ADMIN_BASE_URL = os.getenv("ADMIN_BASE_URL", "https://juliymarian.fly.dev/admin")

app = Flask(__name__)
app.secret_key = os.getenv("FLASK_SECRET_KEY", "cambiame-para-produccion")
# Para que en templates puedas usar {{ config.get('ADMIN_KEY') }}
app.config["ADMIN_KEY"] = ADMIN_KEY

# ---------- DB helpers ----------
def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
    return g.db

@app.teardown_appcontext
def close_db(_exc):
    db = g.pop("db", None)
    if db is not None:
        db.close()

def init_db():
    db = get_db()
    db.execute("""
        CREATE TABLE IF NOT EXISTS rsvps (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT NOT NULL,
            confirma INTEGER NOT NULL,
            menu TEXT,
            mensaje TEXT,
            created_at TEXT NOT NULL
        )
    """)
    db.execute("""
        CREATE TABLE IF NOT EXISTS invitados (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT UNIQUE NOT NULL
        )
    """)
    # --- NUEVA TABLA: GASTOS ---
    db.execute("""
        CREATE TABLE IF NOT EXISTS gastos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            concepto TEXT NOT NULL,
            tipo TEXT NOT NULL CHECK(tipo IN ('por_invitado', 'total')),
            monto REAL NOT NULL,
            notas TEXT,
            created_at TEXT NOT NULL
        )
    """)
    db.commit()

@app.before_request
def ensure_db():
    init_db()

def admin_redirect():
    return redirect(f"{ADMIN_BASE_URL}?key={ADMIN_KEY}")

# ---------- Util menú ----------
ALIASES_VEGGIE = {"veggie", "vegano", "vegan", "vegetariano", "vegetal"}

def normalize_menu(val: str | None) -> str | None:
    if not val:
        return None
    v = val.strip().lower()
    if v in ALIASES_VEGGIE:
        return "veggie"
    if v == "standard":
        return "standard"
    return None

# ---------- Rutas principales (RSVP) ----------
@app.get("/")
def rsvp_form():
    return render_template("rsvp.html")

@app.post("/enviar")
def enviar_rsvp():
    try:
        nombre = (request.form.get("nombre") or "").strip()
        confirma_val = (request.form.get("confirma") or "").strip().lower()
        raw_menu = (request.form.get("menu") or request.form.get("restricciones") or "").strip().lower()
        mensaje = (request.form.get("mensaje") or "").strip()

        errors = []
        if not nombre:
            errors.append("El nombre es obligatorio.")
        if confirma_val not in ("si", "no"):
            errors.append("Indicá si asistís o no.")

        menu = normalize_menu(raw_menu)
        if confirma_val == "si" and menu not in ("standard", "veggie"):
            errors.append("Elegí un menú: Standard o Veggie.")

        db = get_db()
        row_inv = db.execute("SELECT 1 FROM invitados WHERE nombre = ?", (nombre,)).fetchone()
        if not row_inv:
            errors.append("El nombre debe coincidir con un invitado cargado.")

        if errors:
            for e in errors:
                flash(e, "danger")
            return redirect(url_for("rsvp_form"))

        confirma = 1 if confirma_val == "si" else 0
        menu_to_save = menu if confirma == 1 else None

        db.execute("""
            INSERT INTO rsvps (nombre, confirma, menu, mensaje, created_at)
            VALUES (?, ?, ?, ?, ?)
        """, (nombre, confirma, menu_to_save, (mensaje or None),
              datetime.now().isoformat(timespec="seconds")))
        db.commit()

        return redirect(url_for("gracias"), code=303)

    except Exception as ex:
        print("ERROR en /enviar:", ex)
        flash("Ocurrió un error guardando tu confirmación. Probá de nuevo.", "danger")
        return redirect(url_for("rsvp_form"))

@app.get("/gracias")
def gracias():
    return render_template("gracias.html")

@app.get("/admin")
def admin():
    key = request.args.get("key", "")
    if key != ADMIN_KEY:
        abort(401)

    db = get_db()
    rsvps = db.execute("""
        SELECT id, nombre, confirma, menu, mensaje, created_at
        FROM rsvps
        ORDER BY created_at DESC
    """).fetchall()

    invitados = db.execute("""
        SELECT id, nombre
        FROM invitados
        ORDER BY nombre ASC
    """).fetchall()

    total_si = sum(1 for r in rsvps if r["confirma"] == 1)
    total_no = sum(1 for r in rsvps if r["confirma"] == 0)
    total_standard = sum(1 for r in rsvps if r["confirma"] == 1 and (r["menu"] or "").lower() == "standard")
    total_veggie = sum(1 for r in rsvps if r["confirma"] == 1 and (r["menu"] or "").lower() in {"veggie", "vegano"})

    return render_template(
        "admin.html",
        rsvps=rsvps,
        invitados=invitados,
        cant_invitados=len(invitados),
        total_si=total_si,
        total_no=total_no,
        total_standard=total_standard,
        total_veggie=total_veggie,
        key=key
    )

@app.post("/admin/cargar_invitados")
def admin_cargar_invitados():
    key = request.args.get("key", "")
    if key != ADMIN_KEY:
        abort(401)

    lista = (request.form.get("lista") or "").strip()
    if not lista:
        return admin_redirect()

    nombres = []
    for linea in lista.splitlines():
        for nombre in linea.split(","):
            nombre = nombre.strip()
            if nombre:
                nombres.append(nombre)

    db = get_db()
    for n in nombres:
        try:
            db.execute("INSERT OR IGNORE INTO invitados(nombre) VALUES (?)", (n,))
        except:
            pass
    db.commit()
    return admin_redirect()

@app.post("/admin/invitado/update")
def admin_invitado_update():
    key = request.form.get("key", "")
    if key != ADMIN_KEY:
        abort(401)

    inv_id = request.form.get("id")
    nuevo = (request.form.get("nombre") or "").strip()
    cascade = (request.form.get("cascade") == "1")

    if not inv_id or not nuevo:
        flash("Faltan datos para actualizar el invitado.", "danger")
        return admin_redirect()

    db = get_db()
    try:
        row_old = db.execute("SELECT nombre FROM invitados WHERE id = ?", (inv_id,)).fetchone()
        if not row_old:
            flash("Invitado no encontrado.", "danger")
            return admin_redirect()
        viejo = row_old["nombre"]

        db.execute("UPDATE invitados SET nombre = ? WHERE id = ?", (nuevo, inv_id))

        if cascade and viejo != nuevo:
            db.execute("UPDATE rsvps SET nombre = ? WHERE nombre = ?", (nuevo, viejo))

        db.commit()
        flash("Invitado actualizado correctamente.", "success")
    except sqlite3.IntegrityError:
        db.rollback()
        flash("Ya existe un invitado con ese nombre.", "danger")
    except Exception as e:
        db.rollback()
        flash(f"Error al actualizar invitado: {e}", "danger")

    return admin_redirect()

@app.get("/api/invitados")
def api_invitados():
    q = (request.args.get("q") or "").strip()
    if len(q) < 4:
        return jsonify({"ok": True, "items": []})

    db = get_db()
    filas = db.execute(
        """
        SELECT i.nombre
        FROM invitados i
        LEFT JOIN (
           SELECT nombre, MAX(created_at) AS mx
           FROM rsvps
           GROUP BY nombre
        ) ult ON ult.nombre = i.nombre
        LEFT JOIN rsvps r
               ON r.nombre = i.nombre
              AND r.created_at = ult.mx
        WHERE i.nombre LIKE ?
          AND (r.nombre IS NULL OR r.confirma <> 1)
        ORDER BY i.nombre
        LIMIT 5
        """,
        (f"%{q}%",)   # contiene, como tu endpoint viejo
    ).fetchall()

    return jsonify({"ok": True, "items": [f["nombre"] for f in filas]})


@app.post("/admin/rsvp/update")
def admin_rsvp_update():
    key = request.form.get("key", "")
    if key != ADMIN_KEY:
        abort(401)

    db = get_db()

    rid = request.form.get("id")
    nombre = (request.form.get("nombre") or "").strip()
    confirma = request.form.get("confirma")    # "1" o "0"
    raw_menu = (request.form.get("menu") or "").strip().lower()
    mensaje = (request.form.get("mensaje") or "").strip() or None

    if not rid or not nombre or confirma not in ("0", "1"):
        flash("Datos incompletos para actualizar el RSVP.", "danger")
        return admin_redirect()

    menu = normalize_menu(raw_menu)
    if confirma == "0":
        menu = None

    try:
        db.execute(
            """
            UPDATE rsvps
               SET nombre = ?,
                   confirma = ?,
                   menu = ?,
                   mensaje = ?
             WHERE id = ?
            """,
            (nombre, int(confirma), menu, mensaje, rid)
        )
        db.commit()
        flash("RSVP actualizado correctamente.", "success")
    except Exception as e:
        db.rollback()
        flash(f"Error al actualizar: {e}", "danger")

    return admin_redirect()

@app.get("/admin/export.xlsx")
def admin_export_xlsx():
    key = request.args.get("key", "")
    if key != ADMIN_KEY:
        abort(401)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from openpyxl.utils import get_column_letter
    except Exception:
        return Response(
            "Falta instalar openpyxl. Ejecutá:\n\n    pip install openpyxl\n",
            mimetype="text/plain; charset=utf-8",
            status=500,
        )

    db = get_db()
    try:
        rows = db.execute("""
            SELECT i.nombre,
                   r.confirma,
                   r.menu,
                   r.mensaje,
                   r.created_at
            FROM invitados i
            LEFT JOIN (
                SELECT nombre, confirma, menu, mensaje, created_at
                FROM rsvps r1
                WHERE r1.created_at = (
                    SELECT MAX(r2.created_at)
                    FROM rsvps r2
                    WHERE r2.nombre = r1.nombre
                )
            ) r ON r.nombre = i.nombre
            ORDER BY i.nombre
        """).fetchall()
    except Exception as e:
        return Response(f"Error leyendo la base: {e}", mimetype="text/plain; charset=utf-8", status=500)

    respondieron, faltan = [], []
    for row in rows:
        if row["created_at"]:
            confirma_txt = "si" if row["confirma"] == 1 else "no"
            respondieron.append([
                row["nombre"],
                confirma_txt,
                row["menu"] or "",
                row["mensaje"] or "",
                row["created_at"],
            ])
        else:
            faltan.append([row["nombre"]])

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Respondieron"
    headers1 = ["nombre", "confirma", "menu", "mensaje", "fecha_ultima_respuesta"]
    ws1.append(headers1)
    for r in respondieron:
        ws1.append(r)
    bold = Font(bold=True)
    for cell in ws1[1]:
        cell.font = bold
        cell.alignment = Alignment(vertical="center")
    col_widths1 = [30, 10, 14, 50, 24]
    for idx, w in enumerate(col_widths1, start=1):
        ws1.column_dimensions[get_column_letter(idx)].width = w

    ws2 = wb.create_sheet(title="Faltan")
    headers2 = ["nombre"]
    ws2.append(headers2)
    for r in faltan:
        ws2.append(r)
    for cell in ws2[1]:
        cell.font = bold
        cell.alignment = Alignment(vertical="center")
    ws2.column_dimensions["A"].width = 30

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = f"confirmaciones_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    resp = Response(
        bio.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return resp

@app.post("/admin/invitado/delete")
def admin_invitado_delete():
    key = request.form.get("key", "")
    if key != ADMIN_KEY:
        abort(401)

    inv_id = request.form.get("id")
    cascade = request.form.get("cascade_delete") == "1"

    if not inv_id:
        flash("Falta el ID del invitado.", "danger")
        return admin_redirect()

    db = get_db()
    inv = db.execute("SELECT id, nombre FROM invitados WHERE id = ?", (inv_id,)).fetchone()
    if not inv:
        flash("El invitado no existe.", "danger")
        return admin_redirect()

    try:
        if cascade:
            db.execute("DELETE FROM rsvps WHERE nombre = ?", (inv["nombre"],))
        db.execute("DELETE FROM invitados WHERE id = ?", (inv_id,))
        db.commit()
        msg = f'Invitado "{inv["nombre"]}" eliminado'
        msg += " junto con sus confirmaciones." if cascade else "."
        flash(msg, "success")
    except Exception as e:
        db.rollback()
        flash(f"Error al borrar invitado: {e}", "danger")

    return admin_redirect()

# =========================
# ===  MÓDULO GASTOS   ===
# =========================

def parse_monto(s: str) -> float:
    """
    Acepta '12.34', '12,34', '12.345,67', '12345' -> float
    """
    if not s:
        return 0.0
    s = s.strip()
    # Normalización simple: si hay coma y no hay punto, reemplazar coma por punto
    if "," in s and "." not in s:
        s = s.replace(",", ".")
    # Remover separadores de miles comunes
    s = s.replace(" ", "").replace(",", "")
    try:
        return float(s)
    except:
        return float(s.replace(".", "", s.count(".")-1))  # fallback tosco

def get_totales_base(db, n_manual: int | None, base: str):
    # total invitados
    total_invitados = db.execute("SELECT COUNT(*) AS c FROM invitados").fetchone()["c"]

    # total confirmados (según ÚLTIMA respuesta por nombre)
    total_confirmados = db.execute("""
        SELECT COUNT(*) AS c
        FROM (
          SELECT r1.nombre, MAX(r1.created_at) AS mx
          FROM rsvps r1
          GROUP BY r1.nombre
        ) u
        JOIN rsvps r ON r.nombre = u.nombre AND r.created_at = u.mx
        WHERE r.confirma = 1
    """).fetchone()["c"]

    # n_base
    if base == "confirmados":
        n_base = total_confirmados
    elif base == "manual":
        n_base = max(0, int(n_manual or 0))
    else:
        base = "invitados"
        n_base = total_invitados

    return total_invitados, total_confirmados, n_base, base

@app.get("/gastos")
def gastos_panel():
    base = (request.args.get("base") or "invitados").strip().lower()
    try:
        n_manual = int(request.args.get("n") or 0)
    except:
        n_manual = 0

    db = get_db()
    total_invitados, total_confirmados, n_base, base = get_totales_base(db, n_manual, base)

    # Traer gastos y calcular totales en Python
    filas = db.execute("""
        SELECT id, concepto, tipo, monto, notas, created_at
        FROM gastos
        ORDER BY created_at DESC, id DESC
    """).fetchall()

    # Calcular totales
    total_por_invitado = 0.0
    total_totales = 0.0
    filas_calc = []
    for r in filas:
        if r["tipo"] == "por_invitado":
            total_linea = (r["monto"] or 0.0) * n_base
            total_por_invitado += total_linea
        else:
            total_linea = (r["monto"] or 0.0)
            total_totales += total_linea
        filas_calc.append({
            "id": r["id"],
            "concepto": r["concepto"],
            "tipo": r["tipo"],
            "monto": float(r["monto"] or 0.0),
            "notas": r["notas"],
            "created_at": r["created_at"],
            "total_linea": float(total_linea),
        })

    gran_total = total_por_invitado + total_totales
    costo_por_invitado = (gran_total / n_base) if n_base > 0 else 0.0

    return render_template(
        "gastos.html",
        base=base,
        n_manual=n_manual,
        n_base=n_base,
        total_invitados=total_invitados,
        total_confirmados=total_confirmados,
        filas=filas_calc,
        total_por_invitado=total_por_invitado,
        total_totales=total_totales,
        gran_total=gran_total,
        costo_por_invitado=costo_por_invitado
    )

@app.post("/gastos/agregar")
def gastos_agregar():
    base = (request.args.get("base") or "invitados").strip().lower()
    try:
        n_manual = int(request.args.get("n") or 0)
    except:
        n_manual = 0

    concepto = (request.form.get("concepto") or "").strip()
    tipo = (request.form.get("tipo") or "").strip()
    monto_raw = (request.form.get("monto") or "").strip()
    notas = (request.form.get("notas") or "").strip() or None

    if not concepto or tipo not in ("por_invitado", "total"):
        flash("Completá concepto y tipo.", "danger")
        return redirect(url_for("gastos_panel", base=base, n=n_manual))

    monto = parse_monto(monto_raw)
    if monto < 0:
        flash("El monto no puede ser negativo.", "danger")
        return redirect(url_for("gastos_panel", base=base, n=n_manual))

    db = get_db()
    db.execute("""
        INSERT INTO gastos (concepto, tipo, monto, notas, created_at)
        VALUES (?, ?, ?, ?, ?)
    """, (concepto, tipo, monto, notas, datetime.now().isoformat(timespec="seconds")))
    db.commit()
    flash("Gasto agregado.", "success")
    return redirect(url_for("gastos_panel", base=base, n=n_manual))

@app.post("/gastos/editar")
def gastos_editar():
    base = (request.args.get("base") or "invitados").strip().lower()
    try:
        n_manual = int(request.args.get("n") or 0)
    except:
        n_manual = 0

    gid = request.form.get("id")
    concepto = (request.form.get("concepto") or "").strip()
    tipo = (request.form.get("tipo") or "").strip()
    monto_raw = (request.form.get("monto") or "").strip()
    notas = (request.form.get("notas") or "").strip() or None

    if not gid or not concepto or tipo not in ("por_invitado", "total"):
        flash("Datos incompletos para editar.", "danger")
        return redirect(url_for("gastos_panel", base=base, n=n_manual))

    monto = parse_monto(monto_raw)
    if monto < 0:
        flash("El monto no puede ser negativo.", "danger")
        return redirect(url_for("gastos_panel", base=base, n=n_manual))

    db = get_db()
    db.execute("""
        UPDATE gastos
           SET concepto = ?, tipo = ?, monto = ?, notas = ?
         WHERE id = ?
    """, (concepto, tipo, monto, notas, gid))
    db.commit()
    flash("Gasto actualizado.", "success")
    return redirect(url_for("gastos_panel", base=base, n=n_manual))

@app.post("/gastos/borrar/<int:gid>")
def gastos_borrar(gid):
    base = (request.args.get("base") or "invitados").strip().lower()
    try:
        n_manual = int(request.args.get("n") or 0)
    except:
        n_manual = 0

    db = get_db()
    db.execute("DELETE FROM gastos WHERE id = ?", (gid,))
    db.commit()
    flash("Gasto eliminado.", "success")
    return redirect(url_for("gastos_panel", base=base, n=n_manual))

@app.get("/gastos/export.xlsx")
def gastos_export_xlsx():
    base = (request.args.get("base") or "invitados").strip().lower()
    try:
        n_manual = int(request.args.get("n") or 0)
    except:
        n_manual = 0

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from openpyxl.utils import get_column_letter
    except Exception:
        return Response(
            "Falta instalar openpyxl. Ejecutá:\n\n    pip install openpyxl\n",
            mimetype="text/plain; charset=utf-8",
            status=500,
        )

    db = get_db()
    total_invitados, total_confirmados, n_base, base = get_totales_base(db, n_manual, base)

    filas = db.execute("""
        SELECT id, concepto, tipo, monto, notas, created_at
        FROM gastos
        ORDER BY created_at DESC, id DESC
    """).fetchall()

    # armar data + totales
    total_por_invitado = 0.0
    total_totales = 0.0
    rows = []
    for r in filas:
        if r["tipo"] == "por_invitado":
            total_linea = (r["monto"] or 0.0) * n_base
            total_por_invitado += total_linea
        else:
            total_linea = (r["monto"] or 0.0)
            total_totales += total_linea
        rows.append([
            r["created_at"],
            r["concepto"],
            r["tipo"],
            float(r["monto"] or 0.0),
            n_base if r["tipo"] == "por_invitado" else "",
            float(total_linea),
            r["notas"] or "",
        ])

    gran_total = total_por_invitado + total_totales
    costo_por_invitado = (gran_total / n_base) if n_base > 0 else 0.0

    wb = Workbook()

    # Hoja 1: Gastos
    ws1 = wb.active
    ws1.title = "Gastos"
    headers1 = ["fecha", "concepto", "tipo", "monto_base", "n_base", "total_linea", "notas"]
    ws1.append(headers1)
    for row in rows:
        ws1.append(row)

    bold = Font(bold=True)
    for cell in ws1[1]:
        cell.font = bold
        cell.alignment = Alignment(vertical="center")

    widths = [20, 28, 14, 14, 10, 16, 40]
    for i, w in enumerate(widths, start=1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # Hoja 2: Resumen
    ws2 = wb.create_sheet("Resumen")
    ws2.append(["Base usada", n_base])
    ws2.append(["Total por invitado", total_por_invitado])
    ws2.append(["Total (totales)", total_totales])
    ws2.append(["Gran total", gran_total])
    ws2.append(["Costo por invitado", costo_por_invitado])
    for cell in ws2["A1":"A5"][0]:
        cell.font = bold

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = f"gastos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    resp = Response(
        bio.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return resp

# ----------------------

if __name__ == "__main__":
    app.run(debug=True)
